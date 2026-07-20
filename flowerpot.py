"""The main script that implements the tkinter GUI and handles the calculations
of the physical properties the user is interested in."""
import os
import re
import csv
import tkinter as tk
import tkinter.ttk as ttk
from tkinter.filedialog import askopenfilename, askdirectory
import numpy as np
from PIL import Image, ImageTk
from rdkit import Chem
from rdkit.Chem import Descriptors
from openpyxl import Workbook, load_workbook
from utils.mol_list import get_mol_list
from boiled_egg.print_egg import print_eggs
from substructure.filters import check_filters
from log_models.fragment_model import FragmentLogModel
from log_models.log_ml_model import MLPWrapper
from docking.docking import (perform_vina_docking, perform_gold_docking,
custom_gold_docking, custom_vina_docking)
from docking.assay_simulator import build_simulator_window

# Global window variables
FRAME_X_PADDING = 1
FRAME_Y_PADDING = 1
MIN_SIZE_X = 1100
MIN_SIZE_Y = 530
TEXTBOX_WIDTH = 40
TEXTBOX_HEIGHT = 30
IMAGE_FRAME_X = 200
IMAGE_FRAME_Y = 220

# Means and standard deviations of LogS and LogD data for scaling model outputs
LOGS_MEAN = np.array([65.26151553, 60.42754609, 103.15170248])
LOGS_STD = np.array([38.36352282, 55.44868886, 58.93997877])
LOGD_MEAN = np.array([102.87402925, 78.14599844, 158.14487235])
LOGD_STD = np.array([30.67936449, 33.78076728, 47.4308476])

########## FUNCTIONS FOR TKINTER WINDOWS ##########
def get_filename(filename_var, filename_entry):
	"""Implements the 'Browse' button to select an input file."""
	filename = askopenfilename()
	filename_var.set(filename)
	filename_entry.xview_moveto(1)

def get_directory(directory_var, directory_entry):
	"""Implements the 'Browse' button to select a directory."""
	dir_name = askdirectory()
	directory_var.set(dir_name)
	directory_entry.xview_moveto(1)

def activate_print_egg_button():
	"""Ensures that if the 'print egg' option is selected, the permeation
	calculation button will also be selected."""
	global egg_button_value
	global print_egg_button_value
	global print_egg_warning
	if print_egg_button_value.get():
		egg_button_value.set(True)
		print_egg_warning.config(text="Warning: Drawing lots of EGG images "\
			"can take a long time.\nUse small numbers of molecules for better "\
			"performance.", fg="orange")
	else:
		print_egg_warning.config(text="\n")

def activate_egg_button():
	"""Ensures that if the 'Permeation' option is deselected, the 'print egg'
	option will also be deselected."""
	global egg_button_value
	global print_egg_button_value
	global print_egg_warning
	if not egg_button_value.get():
		print_egg_button_value.set(False)
		print_egg_warning.config(text="\n")

def disable_buttons():
	"""Disables the assay simulator and go buttons when calculations are running
	to prevent multiple clicks."""
	global go_button, assay_button
	go_button.config(state=tk.DISABLED)
	assay_button.config(state=tk.DISABLED)
	go_button.update()
	assay_button.update()

def enable_buttons():
	"""Enables the assay simulator and go buttons after the calculation
	function is finished."""
	global go_button, assay_button
	go_button.update()
	assay_button.update()
	go_button.config(state=tk.NORMAL)
	assay_button.config(state=tk.NORMAL)

def on_simulator_close(event):
	"""Releases the assay and go buttons once the simulator window is closed."""
	global simulator_window
	global assay_button
	if event.widget != simulator_window:
		return
	enable_buttons()

def run_assay_simulator():
	"""Starts the assay simulator pop-up window if docking scores calculations
	have been performed and the data is available in the global 'properties'
	dictionary."""
	global properties
	global protein_selection
	global window, mols, file_names
	docking_score_keys = tuple(score_key for score_key in [protein + program \
		for protein in list(protein_selection.keys()) + ["Custom"] \
		for program in (" GOLD score", " Vina score")] \
		if properties.get(score_key) is not None)
	if len(properties) == 0 or len(docking_score_keys) == 0:
		return
	global simulator_window
	simulator_window = build_simulator_window(window, mols, file_names,
		properties, docking_score_keys)
	simulator_window.bind("<Destroy>", on_simulator_close)
	# Prevent multiple simulator windows from opening
	disable_buttons()

def activate_docking_warning():
	"""Highlights a warning message to the user about the computational cost of
	docking when a relevant option is selected."""
	global gold_button_value
	global vina_button_value
	global custom_gold_button_value
	global custom_vina_button_value
	if gold_button_value.get() or vina_button_value.get() \
	or custom_gold_button_value.get() or custom_vina_button_value.get():
		docking_warning.config(text="Warning: Molecular docking is "\
			"computationally intensive.\nOnly use for a small number of "\
			"molecules at a time.", fg="red")
	else:
		docking_warning.config(text="\n")

def backward():
	"""Implements the functioning of the previous button for the pop-up 
	image window."""
	global mol_no
	global file_names
	global smiles_heading
	global image_list
	global picture_frame
	global image_labels
	mol_no = (mol_no - 1) % len(file_names)
	smiles_heading.configure(text=file_names[mol_no])
	if len(image_list[mol_no]) > 0:
		picture_frame.grid_propagate(True)
		for i in range(len(image_labels)):
			if i < len(image_list[mol_no]):
				image_labels[i].configure(image=image_list[mol_no][i],
					text=None)
			else:
				image_labels[i].configure(image="", text=None)
		picture_frame.configure(height=picture_frame["height"],
			width=picture_frame["width"])
	else:
		picture_frame.grid_propagate(False)
		image_labels[0].configure(image="", text="No images for this molecule.")
		picture_frame.configure(width=IMAGE_FRAME_X, height=IMAGE_FRAME_Y)
		for label in image_labels[1:]:
			label.configure(image="", text=None)

def forward():
	"""Implements the functioning of the next button for the pop-up image
	window."""
	global mol_no
	global file_names
	global smiles_heading
	global image_list
	global picture_frame
	global image_labels
	mol_no = (mol_no + 1) % len(file_names)
	smiles_heading.configure(text=file_names[mol_no])
	if len(image_list[mol_no]) > 0:
		picture_frame.grid_propagate(True)
		for i in range(len(image_labels)):
			if i < len(image_list[mol_no]):
				image_labels[i].configure(image=image_list[mol_no][i],
					text=None)
			else:
				image_labels[i].configure(image="", text=None)
		picture_frame.configure(height=picture_frame["height"],
			width=picture_frame["width"])
	else:
		picture_frame.grid_propagate(False)
		image_labels[0].configure(image="", text="No images for this molecule.")
		picture_frame.configure(width=IMAGE_FRAME_X, height=IMAGE_FRAME_Y)
		for label in image_labels[1:]:
			label.configure(image="", text=None)

def make_popup_image_window(file_names, egg_image_files, filter_image_files):
	"""
	Creates the pop-up window to display the HARDBOILED-EGG and substructure
	filter images.
	Args:
		file_names: list of str, the base names for the files that the data for
			each molecule may be saved to.
		egg_image_files: list of str, the paths to the egg image files
		filter_image_files: list of list of str, the paths to the substructure
			filter images for each filter applied to each molecule
	"""
	# Step 1. Read egg and filter images
	global mol_no
	global print_egg_button_value
	global image_list
	global filter_button_values
	mol_no = 0
	image_list = [list() for i in range(len(file_names))]
	if print_egg_button_value.get() and egg_image_files is not None \
	and len(egg_image_files) == len(file_names):
		for i, egg_image in enumerate(egg_image_files):
			image_list[i].append(ImageTk.PhotoImage(Image.open(egg_image)))
	if any(filt.get() for filt in filter_button_values.values()) \
	and filter_image_files is not None and len(filter_image_files) > 0 \
	and len(filter_image_files[0]) == len(file_names):
		for filter_image_list in filter_image_files:
			for i, filter_image in enumerate(filter_image_list):
				if filter_image is not None:
					image_list[i].append(ImageTk.PhotoImage(Image.open(
						filter_image)))
	# Step 2. Build the pop-up window
	global window
	global picture_frame
	global smiles_heading
	global image_labels
	image_window = tk.Toplevel(window)
	image_window.resizable(False, False)
	image_window.title("EGG and Filter Images")
	picture_frame = tk.Frame(image_window)
	picture_frame.grid(row=0, column=0, columnspan=2, rowspan=2, sticky="nsew")
	smiles_heading = tk.Label(picture_frame, text=file_names[mol_no])
	smiles_heading.grid(row=0, column=0, columnspan=2, sticky="nsew")
	image_labels = list()
	for i in range(len(filter_button_values.keys()) + 1):
		label = tk.Label(picture_frame)
		label.grid(row=1, column=i, sticky="nsew")
		image_labels.append(label)
	if len(image_list[mol_no]) > 0:
		picture_frame.grid_propagate(True)
		for i, label in enumerate(image_labels):
			if i < len(image_list[mol_no]):
				label.configure(image=image_list[mol_no][i], text=None)
			else:
				label.configure(image="", text=None)
		picture_frame.configure(height=picture_frame["height"],
			width=picture_frame["width"])
	else:
		picture_frame.grid_propagate(False)
		image_labels[0].configure(image="", text="No images for this molecule.")
		picture_frame.configure(width=IMAGE_FRAME_X, height=IMAGE_FRAME_Y)
		for label in image_labels[1:]:
			label.configure(image="", text=None)
	button_frame = tk.Frame(image_window)
	button_frame.grid(row=2, column=0, columnspan=2, sticky="nsw")
	backward_button = tk.Button(button_frame, text="Previous", command=backward)
	backward_button.grid(row=2, column=0, sticky="nsew")
	forward_button = tk.Button(button_frame, text="Next", command=forward)
	forward_button.grid(row=2, column=1, sticky="nsew")

########## END FUNCTIONS FOR TKINTER WINDOWS ##########

########## FUNCTIONS FOR DATA AND FILES ##########
def check_file_readable(filename):
	"""
	Checks if a given file name exists and can be read.
	Args:
		filename: str, the name of the file to be checked.
	Returns:
		bool: check on whether the file can be opened and read.
	"""
	if not os.path.exists(filename):
		print("ERROR: File '%s' does not exist." % filename)
		return False
	elif not os.access(filename, os.R_OK):
		print("ERROR: File '%s' could not be read - permission denied." \
			% filename)
		return False
	return True

def check_file_writeable(filename):
	"""
	Checks if a file can be written to.
	Args:
		filename: str, the name of the file to be written to.
	Returns:
		bool: check on whether the file can be written to
	"""
	directory = os.path.dirname(filename)
	directory = os.getcwd() if directory == "" else directory
	if not os.access(directory, os.W_OK):
		print("ERROR: File '%s' could not be written to." % filename)
		return False
	return True

def read_input_from_textbox():
	"""
	Reads the smiles strings from the input text box and returns all
	non-empty entries and 'empty' names.
	Returns:
		smiles_strings: list of str, all non-empty strings in the input text box
			(may not actually be valid smiles strings).
		names: list of None, the molecules in the input text box are not
			associated with name strings.
	"""
	global smiles_text
	smiles_strings = list()
	names = list()
	smiles_input = smiles_text.get("1.0", "end-1c")
	for smiles in smiles_input.splitlines():
		if smiles != "":
			smiles_strings.append(smiles)
			names.append(None)
	return smiles_strings, names

def read_from_csv_file(filename):
	"""
	Reads input smiles strings and associated molecule names from a CSV file.
	Args:
		filename: str, the CSV filename to read.
	Returns:
		smiles_strings: list of str, the smiles strings of the molecules in the
			file. May not be actual, valid smiles.
		names: list of str, the names that each of the molecules are associated
			if they can be found.
	"""
	smiles_strings = list()
	names = list()
	file = open(filename, "r")
	csv_reader = csv.reader(file)
	first_line = next(csv_reader)
	lower_first_line = [item.lower() for item in first_line]
	smiles_pos = None
	name_pos = None
	# Check if the file has headings and find the smiles and name positions
	if "smiles" in lower_first_line:
		smiles_pos = lower_first_line.index("smiles")
		if "name" in lower_first_line:
			name_pos = lower_first_line.index("name")
		elif "names" in lower_first_line:
			name_pos = lower_first_line.index("names")
	else:
		# Try looking for a valid smiles string in the top line instead
		print("Could not find 'smiles' in column heading, looking for a valid "\
			"SMILES string the first row instead...")
		for i, item in enumerate(first_line):
			mol = Chem.MolFromSmiles(item, sanitize=False)
			if mol is not None and item != "":
				smiles_pos = i
				smiles_strings.append(item)
				names.append(None)
				break
		if smiles_pos is None:
			print("ERROR: Failed to find valid SMILES string in first line "\
				"of file %s." % filename)
	# Now read all of the smiles (and names) from the file
	if smiles_pos is not None:
		for line in csv_reader:
			if smiles_pos < len(line):
				smiles_strings.append(line[smiles_pos])
			else:
				print("WARNING: Line was shorter than expected position of "\
					"SMILES string.")
				continue
			if name_pos is not None:
				if name_pos < len(line):
					names.append(line[name_pos])
				else:
					print("WARNING: Line was shorter than expected position "\
						"of molecule name.")
					names.append(None)
			else:
				names.append(None)
	file.close()
	return smiles_strings, names

def read_input_from_xlsx_file(filename):
	"""
	Reads input smiles strings and associated molecule names from an Excel file.
	Args:
		filename: str, the Excel .xlsx filename to read.
	Returns:
		smiles_strings: list of str, the smiles strings of the molecules in the
			file. May not be actual, valid smiles.
		names: list of str, the names that each of the molecules are associated
			if they can be found.
	"""
	smiles_strings = list()
	names = list()
	workbook = load_workbook(filename, read_only=True)
	worksheet = workbook[workbook.sheetnames[0]]
	smiles_pos = None
	name_pos = None
	rows = list(worksheet.rows)
	# Get the first row of the spreadsheet as a list of strings
	first_line = [str(cell.value) for cell in rows[0]]
	lower_first_line = [item.lower() for item in first_line]
	if "smiles" in lower_first_line:
		smiles_pos = lower_first_line.index("smiles")
		if "name" in lower_first_line:
			name_pos = lower_first_line.index("name")
		elif "names" in lower_first_line:
			name_pos = lower_first_line.index("names")
	else:
		print("Could not find 'smiles' column in heading, looking for valid "\
			"SMILES string in first row instead...")
		for i, item in enumerate(first_line):
			mol = Chem.MolFromSmiles(item, sanitize=False)
			if mol is not None:
				smiles_pos = i
				smiles_strings.append(item)
				names.append(None)
				break
		if smiles_pos is None:
			print("ERROR: Failed to find valid SMILES string in first line of "\
				"file '%s'." % filename)
	# Now read all of the smiles (and names) from the file
	if smiles_pos is not None:
		for row in rows[1:]:
			if smiles_pos < len(row):
				smiles_strings.append(row[smiles_pos].value)
			else:
				print("WARNING: Line was shorter than expected position of "\
					"SMILES string.")
				continue
			if name_pow is not None:
				if name_pos < len(row):
					names.append(row[name_pow].value)
				else:
					print("WARNING: Line was shorter than expected position "\
						"of molecule name.")
					names.append(None)
			else:
				names.append(None)
	workbook.close()
	return smiles_strings, names

def make_molecule_filenames(names):
	"""
	Associates each of the molecule names with a base name for the files its
	data may be saved to. Unnamed molecules will be given the generic name
	'moleculeN' with N increasing in order.
	Args:
		names: list of str or None, the provided names of each of the molecules.
			Empty or NoneType names will be given a generic name for their file.
	Returns:
		file_names: list of str, the base names for the files that the data for
			each molecule may be saved to.
	"""
	file_names = list()
	name_counts = dict()
	n = 1
	for name in names:
		if name is None or name == "" or name.isspace():
			file_names.append("molecule" + str(n))
			n += 1
		else:
			# Make sure there are no operating system invalid characters in name
			valid_name = re.sub("[\\:/]", "", name)
			# Ensure that the same name does not appear twice
			if name_counts.get(valid_name) is None:
				name_counts[valid_name] = 0
			else:
				name_counts[valid_name] += 1
				valid_name += "-" + str(name_counts[valid_name])
			file_names.append(valid_name)
	return file_names

def write_data_csv_textbox(filename, smiles_strings, file_names, error_messages,
properties, prop_key_list):
	"""
	Writes all collected data to a CSV file and the output textbox at the same
	time.
	Args:
		filename: str, path to the CSV file that will be written.
		smiles_strings: list of str, the SMILES strings for each molecule.
		file_names: list of str, the validated names for each of the molecules.
		error_messages: list of str, the error messages for the molecules.
		properties: dict of list of float, dictionary containing the properties
			of the molecules.
		prop_key_list: list of str, the keys of the properties dictionary in the
			order they were added.
	"""
	global output_text
	global quiet_errors_button_value
	output_file = open(filename, "w")
	heading = "names,smiles," \
		+ ','.join(prop_key for prop_key in prop_key_list) + '\n'
	output_text.insert(tk.END, heading)
	output_file.write(heading)
	m = 0
	for smiles, err_message in zip(smiles_strings, error_messages):
		if err_message is not None:
			if type(err_message) == str and not quiet_errors_button_value.get():
				err_message += '\n'
				output_text.insert(tk.END, err_message)
				output_file.write(err_message)
			continue
		line = [file_names[m], smiles]
		for prop_key in prop_key_list:
			prop = properties[prop_key][m]
			if type(prop) == float:
				line.append("%.2f" % prop)
			else:
				line.append(str(prop))
		line = ','.join(line) + '\n'
		output_text.insert(tk.END, line)
		output_file.write(line)
		m += 1
	output_text.configure(state="disabled")
	output_file.close()

def write_data_xlsx_textbox(filename, smiles_strings, file_names,
error_messages, properties, prop_key_list):
	"""
	Writes all collected data to a Excel file and the output textbox at the same
	time.
	Args:
		filename: str, path to the Excel file that will be written.
		smiles_strings: list of str, the SMILES strings for each molecule.
		file_names: list of str, the validated names for each of the molecules.
		error_messages: list of str, the error messages for the molecules.
		properties: dict of list of float, dictionary containing the properties
			of the molecules.
		prop_key_list: list of str, the keys of the properties dictionary in the
			order they were added.
	"""
	global output_text
	global quiet_errors_button_value
	output_workbook = Workbook(write_only=True)
	output_file = output_workbook.create_sheet("Sheet1")
	heading = ["names", "smiles"] + prop_key_list
	output_file.append(heading)
	heading = ','.join(prop_key for prop_key in prop_key_list) + '\n'
	output_text.insert(tk.END, heading)
	m = 0
	for smiles, err_message in zip(smiles_strings, error_messages):
		if err_message is not None:
			if type(err_message) == str and not quiet_errors_button_value.get():
				output_file.append([err_message])
				output_text.insert(tk.END, err_message + '\n')
			continue
		line = [file_names[m], smiles]
		for prop_key in prop_key_list:
			prop = properties[prop_key][m]
			if type(prop) == float:
				line.append("%.2f" % prop)
			else:
				line.append(str(prop))
		output_file.append(line)
		line = ','.join(line) + '\n'
		output_text.insert(tk.END, line)
		m += 1
	output_text.configure(state="disabled")
	output_workbook.save(filename)
	output_workbook.close()

########## END FUNCTIONS FOR DATA AND FILES ##########

########## FUNCTIONS FOR MOLECULE PROPERTIES ##########
def calc_numerical_properties(mols, properties, prop_key_list):
	"""
	Calculates the numerical properties for each of the molecules ('Log'
	properties and topological polar surface area).
	Args:
		mols: list of RDKit mol objects, list of the molecules for which
			properties are to be calculated.
		properties: dict of list of float, dictionary containing the properties
			of the molecules.
		prop_key_list: list of str, the keys of the properties dictionary in the
			order they were added.
	"""
	global numprop_button_values
	global log_predictor_button_value
	for prop, wanted in numprop_button_values.items():
		if not wanted.get():
			continue
		if prop == "LogD":
			if log_predictor_button_value.get() == 0:
				logd_model = FragmentLogModel("log_models/Crippen.txt")
				logd_model.load_coef_file("log_models/lipo_coefs.npy")
			elif log_predictor_button_value.get() == 1:
				logd_weights = np.load("log_models/logd_network_weights.npz")
				logd_biases = np.load("log_models/logd_network_biases.npz")
				logd_model = MLPWrapper(logd_weights, logd_biases, LOGD_MEAN,
					LOGD_STD)
			predictions = logd_model.predict(mols)
			properties[prop] = list(float(pred) for pred in predictions)
		elif prop == "LogS":
			if log_predictor_button_value.get() == 0:
				logs_model = FragmentLogModel("log_models/Crippen.txt")
				logs_model.load_coef_file("log_models/sol_coefs.npy")
			elif log_predictor_button_value.get() == 1:
				logs_weights = np.load("log_models/logs_network_weights.npz")
				logs_biases = np.load("log_models/logs_network_biases.npz")
				logs_model = MLPWrapper(logs_weights, logs_biases, LOGS_MEAN,
					LOGS_STD)
			predictions = logs_model.predict(mols)
			properties[prop] = list(float(pred) for pred in predictions)
		elif prop == "LogP":
			properties[prop] = list()
			for mol in mols:
				properties[prop].append(Descriptors.MolLogP(mol))
		elif prop == "tPSA":
			properties[prop] = list()
			for mol in mols:
				properties[prop].append(Descriptors.TPSA(mol))
		prop_key_list.append(prop)

def calc_hardboiled_egg(mols, properties, prop_key_list, file_names):
	"""
	Runs the HARDBOILED-EGG calculations for the molecules.
	Args:
		mols: list of RDKit mol objects, list of the molecules for which
			properties are to be calculated.
		properties: dict of list of float, dictionary containing the properties
			of the molecules.
		prop_key_list: list of str, the keys of the properties dictionary in the
			order they were added.
		file_names: list of str, the base names for the files that the data for
			each molecule may be saved to.
	Returns:
		egg_image_files: list of str, the paths to the drawn HARDBOILED-EGG
			images.
	"""
	global print_egg_button_value
	egg_image_files, hia_perms, bbb_perms = print_eggs(mols, file_names,
		print_egg_button_value.get())
	prop_key_list.append("HIA permeation")
	prop_key_list.append("BBB permeation")
	properties["HIA permeation"] = hia_perms
	properties["BBB permeation"] = bbb_perms
	return egg_image_files

def calc_substructure_filters(mols, properties, prop_key_list, file_names):
	"""
	Performs substructure matching on the molecules.
	Args:
		mols: list of RDKit mol objects, list of the molecules for which
			properties are to be calculated.
		properties: dict of list of float, dictionary containing the properties
			of the molecules.
		prop_key_list: list of str, the keys of the properties dictionary in the
			order they were added.
		file_names: list of str, the base names for the files that the data for
			each molecule may be saved to.
	Returns:
		filter_image_files: list of list of str, the paths to the substructure
			filter images for each set of filters applied to each molecule
	"""
	global filter_button_values
	filter_image_files = list()
	for filt in sorted(filter_button_values.keys()):
		if not filter_button_values[filt].get():
			continue
		image_files, hit_counts = check_filters(mols, file_names, filt,
			print_images=True)
		filter_image_files.append(image_files)
		prop_key = filt + " hits"
		prop_key_list.append(prop_key)
		properties[prop_key] = hit_counts
	return filter_image_files

def calc_builtin_docking(mols, properties, prop_key_list, file_names):
	"""
	Performs docking calculations on the built-in proteins with the molecules
	using GOLD and Autodock Vina.
	Args:
		mols: list of RDKit mol objects, list of the molecules for which
			properties are to be calculated.
		properties: dict of list of float, dictionary containing the properties
			of the molecules.
		prop_key_list: list of str, the keys of the properties dictionary in the
			order they were added.
		file_names: list of str, the base names for the files that the data for
			each molecule may be saved to.
	"""
	global protein_selection
	global gold_button_value
	global vina_button_value
	global gold_install_path
	for protein in protein_selection.keys():
		if not protein_selection[protein].get():
			continue
		if gold_button_value.get():
			docking_scores = perform_gold_docking(mols, file_names, protein,
				gold_install_path.get())
			prop_key = protein + " GOLD score"
			if docking_scores is None:
				print("ERROR: Could not perform docking score calculations "
					"for protein: %s with GOLD." % protein)
			else:
				prop_key_list.append(prop_key)
				properties[prop_key] = docking_scores
		if vina_button_value.get():
			docking_scores = perform_vina_docking(mols, file_names, protein)
			prop_key = protein + " Vina score"
			if docking_scores is None:
				print("ERROR: Could not perform docking score calculations "
					"for protein: %s with Vina." % protein)
			else:
				prop_key_list.append(prop_key)
				properties[prop_key] = docking_scores

def calc_custom_gold_docking(mols, properties, prop_key_list, file_names):
	"""
	Performs docking calculations with GOLD on a user-supplied protein and their
	custom configuration.
	Args:
		mols: list of RDKit mol objects, list of the molecules for which
			properties are to be calculated.
		properties: dict of list of float, dictionary containing the properties
			of the molecules.
		prop_key_list: list of str, the keys of the properties dictionary in the
			order they were added.
		file_names: list of str, the base names for the files that the data for
			each molecule may be saved to.
	"""
	global conf_filename
	global gold_install_path
	custom_gold_scores = custom_gold_docking(mols, file_names,
		conf_filename.get(), gold_install_path.get())
	if custom_gold_scores is None:
		print("ERROR: Could not perform custom docking score calculations "\
			"with GOLD.")
		return
	prop_key = "Custom GOLD score"
	prop_key_list.append(prop_key)
	properties[prop_key] = custom_gold_scores

def calc_custom_vina_docking(mols, properties, prop_key_list, file_names):
	"""
	Performs docking calculations with Autodock Vina on a user-supplied protein
	and their custom configuration.
	Args:
		mols: list of RDKit mol objects, list of the molecules for which
			properties are to be calculated.
		properties: dict of list of float, dictionary containing the properties
			of the molecules.
		prop_key_list: list of str, the keys of the properties dictionary in the
			order they were added.
		file_names: list of str, the base names for the files that the data for
			each molecule may be saved to.
	"""
	global pdbqt_filename
	global config_filename
	custom_vina_scores = custom_vina_docking(mols, file_names,
		pdbqt_filename.get(), config_filename.get())
	if custom_vina_scores is None:
		print("ERROR: Could not perform custom docking score calculcations "\
			"with Vina.")
		return
	prop_key = "Custom Vina score"
	prop_key_list.append(prop_key)
	properties[prop_key] = custom_vina_scores

########## END FUNCTIONS FOR MOLECULE PROPERTIES ##########

def handle_calculate_button():
	"""
	Handles all inputs and outputs to the program, as well as the calculations
	of all physical properties that selected using the checkbutton options.

	The first step is to read in any SMILES strings in the
	program's input text box, provided the button that states this input
	box is to be ignored is not checked.

	If provided with an existing spreadsheet file name, the program will then
	attempt to find a column titled 'smiles' from which to read the molecule
	SMILES strings. Failing to find a column heading named 'smiles' (case-
	insensitive), the program will then look for the first column in the first
	row of the spreadsheet that contains a valid SMILES string and will use this
	column to read in the SMILES strings.

	Then, the program will perform all calculations that were requested by the
	user using the check boxes. First LogD, LogP, LogS and TPSA will be
	calculated for all of the input molecules that had valid SMILES strings. The
	user requested model type (Crippen linear model or neural network) will be
	used to calculate LogD and LogS.

	Next, the estimation of intestinal absorption and blood-brain barrier
	permeation using the HARDBOILED-EGG model will be performed, the
	substructure filters will be applied to all of the input molecules and then
	a pop-up window will be created that will display the visualisations of
	these results.

	The last optional property calculations that will be performed are 
	molecular docking simulations using either of the programs Autodock Vina or
	GOLD, if the executable binaries can be located. The docking scores may also
	be used in pedagogical simulations of assays for student's molecules using
	the 'assay simulator' pop-up window. The user also has the option to provide
	customised Vina and GOLD job files, which will be run after the jobs
	specifying an in-built protein target.

	The program will write all of the results to the output text box on the
	right hand side of the GUI. Unless it is requested by using a check button
	that errors from incorrect SMILES strings are ignored, then error messages
	for these strings will also be written to the output text box.

	Finally, the program will write all of the calculated results to an output
	file, if a file name is provided. If requested as a .xlsx file, the results
	will be written to an Excel spreadsheet on a worksheet named 'Sheet1'. If
	the file extension is not one of '.csv' or '.xlsx', the .csv file extension
	will be added to the provided name and results will be written in the csv
	format.
	"""
	global go_button, assay_button
	# Disable the buttons so multiple clicks won't be registered
	disable_buttons()

	global properties
	properties = dict()

	global output_text
	global ignore_textbox_button_value
	
	# Delete output text when the job has started
	output_text.configure(state="normal")
	output_text.delete("1.0", "end")
	output_text.update_idletasks()

	smiles_strings = list()
	names = list()
	if not ignore_textbox_button_value.get():
		read_smiles_strings, read_names = read_input_from_textbox()
		smiles_strings.extend(read_smiles_strings)
		names.extend(read_names)

	input_filename = csv_filename.get()
	if input_filename != "" and not check_file_readable(input_filename):
		print("ERROR: Input file '%s' could not be read." % input_filename)
	if input_filename.endswith(".csv"):
		read_smiles_strings, read_names = read_from_csv_file(input_filename)
		smiles_strings.extend(read_smiles_strings)
		names.extend(read_names)
	elif input_filename.endswith(".xlsx"):
		read_smiles_strings, read_names = read_from_xlsx_file(input_filename)
		smiles_strings.extend(read_smiles_strings)
		names.extend(read_names)
	elif input_filename != "":
		print("ERROR: Unsupported input file type, please use either a .csv "\
			"file or a .xlsx spreadsheet.")

	global mols
	mols, errs, names = get_mol_list(smiles_strings, names=names)
	if len(mols) == 0:
		print("ERROR: No valid SMILES strings were found.")
		enable_buttons()
		return

	global file_names
	file_names = make_molecule_filenames(names)

	prop_key_list = list()
	egg_image_files = list()
	filter_image_files = list()

	global numprop_button_values
	global egg_button_value
	global filter_button_values
	global gold_button_value
	global vina_button_value
	global custom_gold_button_value
	global custom_vina_button_value
	global conf_filename
	global pdbqt_filename
	global config_filename
	if any(wanted.get() for wanted in numprop_button_values.values()):
		calc_numerical_properties(mols, properties, prop_key_list)
	if egg_button_value.get():
		egg_image_files = calc_hardboiled_egg(mols, properties, prop_key_list,
			file_names)
	if any(filt.get() for filt in filter_button_values.values()):
		filter_image_files = calc_substructure_filters(mols, properties,
			prop_key_list, file_names)
	if gold_button_value.get() or vina_button_value.get():
		calc_builtin_docking(mols, properties, prop_key_list, file_names)
	if custom_gold_button_value.get():
		if check_file_readable(conf_filename.get()):
			calc_custom_gold_docking(mols, properties, prop_key_list,
				file_names)
		else:
			print("ERROR: Could not perform custom docking score "\
				"calculations with GOLD.")
	if custom_vina_button_value.get():
		if check_file_readable(pdbqt_filename.get()) \
		and check_file_readable(config_filename.get()):
			calc_custom_vina_docking(mols, properties, prop_key_list,
				file_names)
		else:
			print("ERROR: Could not perform custom docking score "\
				"calculations with Autodock Vina.")

	global print_egg_button_value
	if print_egg_button_value.get() \
	or any(filt.get() for filt in filter_button_values.values()):
		make_popup_image_window(file_names, egg_image_files, filter_image_files)

	global output_foldername
	global output_filename
	out_foldername = output_foldername.get()
	out_filename = output_filename.get()
	if out_filename == "":
		out_filename = "output.csv"
	elif not out_filename.endswith(".csv") \
	and not out_filename.endswith(".xlsx"):
		out_filename += ".csv"
		print("WARNING: Output file name does not end with either '.xlsx' or "\
			"'.csv', defaulting to a CSV file '%s'." % out_filename)
	out_filename = os.path.join(out_foldername, out_filename)
	if not check_file_writeable(out_filename):
		enable_buttons()
		return
	print("Writing data to %s..." % out_filename)
	if out_filename.endswith(".csv"):
		write_data_csv_textbox(out_filename, smiles_strings, file_names, errs,
			properties, prop_key_list)
	elif out_filename.endswith(".xlsx"):
		write_data_xlsx_textbox(out_filename, smiles_strings, file_names, errs,
			properties, prop_key_list)

	# Enable the buttons once done
	enable_buttons()


if __name__ == "__main__":
	properties = dict()

	window = tk.Tk()
	window.title("Flower Pot")
	window.minsize(MIN_SIZE_X, MIN_SIZE_Y)

	# SMILES input textbox on left of GUI
	smiles_frame = tk.Frame()
	smiles_frame.grid(row=0, column=0, sticky="nsew", padx=FRAME_X_PADDING,
		pady=FRAME_Y_PADDING)
	smiles_label = ttk.Label(smiles_frame, text="SMILES Input:")
	smiles_label.grid(row=0, column=0, sticky="w")
	smiles_text = tk.Text(smiles_frame, width=TEXTBOX_WIDTH,
		height=TEXTBOX_HEIGHT)
	smiles_text.grid(row=1, column=0, sticky="nsew")
	smiles_frame.grid_columnconfigure(0, weight=1)
	smiles_frame.grid_rowconfigure(1, weight=1)

	# Configuration options in middle of GUI
	option_frame = tk.Frame()
	option_frame.grid(row=0, column=1, sticky="nw", padx=FRAME_X_PADDING,
		pady=FRAME_Y_PADDING)
	# Input and output file options
	input_label = ttk.Label(option_frame, text="Input File Path:")
	input_label.grid(row=0, column=0, sticky="nw")
	csv_filename = tk.StringVar()
	csv_entry = ttk.Entry(option_frame, textvariable=csv_filename,
		state="disabled")
	csv_entry.grid(row=1, column=0, sticky="nw")
	browse_button = ttk.Button(option_frame, text="Browse",
		command=lambda: get_filename(csv_filename, csv_entry))
	browse_button.grid(row=1, column=1, sticky="nw")
	ignore_textbox_button_value = tk.BooleanVar()
	ignore_textbox_button = ttk.Checkbutton(option_frame,
		variable=ignore_textbox_button_value, text="Only use input from file")
	ignore_textbox_button.grid(row=2, column=0, sticky="nw")
	quiet_errors_button_value = tk.BooleanVar()
	quiet_errors_button = ttk.Checkbutton(option_frame,
		variable=quiet_errors_button_value,
		text="Quieten invalid SMILES errors")
	quiet_errors_button.grid(row=3, column=0, sticky="nw")
	output_folder_label = ttk.Label(option_frame, text="Output Folder:")
	output_folder_label.grid(row=4, column=0, sticky="nw")
	output_file_label = ttk.Label(option_frame, text="Output File Name:")
	output_file_label.grid(row=4, column=2, sticky="nw")
	output_foldername = tk.StringVar()
	output_folder_entry = ttk.Entry(option_frame,
		textvariable=output_foldername, state="disabled")
	output_folder_entry.grid(row=5, column=0, sticky="nw")
	output_browse_button = ttk.Button(option_frame, text="Select Folder",
		command=lambda: get_directory(output_foldername, output_folder_entry))
	output_browse_button.grid(row=5, column=1, sticky="nw")
	output_filename = tk.StringVar()
	output_entry = ttk.Entry(option_frame, textvariable=output_filename)
	output_entry.grid(row=5, column=2, sticky="nw")

	# Numerical properties section
	property_label = ttk.Label(option_frame, text="Numerical Properties:")
	property_label.grid(row=6, column=0, sticky="nw")
	numprop_button_values = {"LogD": tk.BooleanVar(), "LogS": tk.BooleanVar(),
		"LogP": tk.BooleanVar(), "tPSA": tk.BooleanVar()}
	for i, num_prop in enumerate(("LogD", "LogS", "LogP", "tPSA")):
		checkbox = ttk.Checkbutton(option_frame,
			variable=numprop_button_values[num_prop], text=num_prop)
		checkbox.grid(row=7, column=i, sticky="w")
	log_predictor_label = ttk.Label(option_frame, text="LogD/LogS Predictor:")
	log_predictor_label.grid(row=8, column=0, sticky="nw")
	log_predictor_button_value = tk.IntVar()
	crippen_button = ttk.Checkbutton(option_frame,
		variable=log_predictor_button_value, onvalue=0, offvalue=0,
		text="Crippen Linear Model")
	ml_button = ttk.Checkbutton(option_frame,
		variable=log_predictor_button_value, onvalue=1, offvalue=1,
		text="Machine Learning")
	crippen_button.grid(row=9, column=0, sticky="nw")
	ml_button.grid(row=9, column=1, sticky="nw")

	# HARDBOILED-EGG section
	egg_label = ttk.Label(option_frame, text="HARDBOILED-EGG:")
	egg_label.grid(row=10, column=0, sticky="w")
	egg_button_value = tk.BooleanVar()
	print_egg_button_value = tk.BooleanVar()
	egg_button = ttk.Checkbutton(option_frame, variable=egg_button_value,
		text="Permeation", command=activate_egg_button)
	egg_button.grid(row=11, column=0, sticky="w")
	print_egg_button = ttk.Checkbutton(option_frame,
		variable=print_egg_button_value, text="Print EGG",
		command=activate_print_egg_button)
	print_egg_button.grid(row=11, column=1, sticky="w")
	print_egg_warning = tk.Label(option_frame, text="\n")
	print_egg_warning.grid(row=12, column=0, columnspan=3, sticky="nsew")

	# Substructure filters section
	filter_label = ttk.Label(option_frame, text="Substructure Filters:")
	filter_label.grid(row=13, column=0, sticky="w")
	filter_button_values = {"PAINS_A": tk.BooleanVar(),
		"PAINS_B": tk.BooleanVar(), "PAINS_C": tk.BooleanVar(),
		"BRENK": tk.BooleanVar(), "NIH": tk.BooleanVar()}
	for i, filt in enumerate(sorted(list(filter_button_values.keys()))):
		checkbox = ttk.Checkbutton(option_frame,
			variable=filter_button_values[filt], text=filt)
		checkbox.grid(row=14, column=i, sticky="w")

	# Docking section
	docking_label = ttk.Label(option_frame, text="Molecular Docking Program:")
	docking_label.grid(row=15, column=0, sticky="w")
	docking_warning = tk.Label(option_frame, text="\n")
	docking_warning.grid(row=16, column=0, columnspan=3, sticky="nsew")
	gold_button_value = tk.BooleanVar()
	vina_button_value = tk.BooleanVar()
	gold_button = ttk.Checkbutton(option_frame, variable=gold_button_value,
		text="GOLD", command=activate_docking_warning)
	vina_button = ttk.Checkbutton(option_frame, variable=vina_button_value,
		text="AutoDock Vina", command=activate_docking_warning)
	gold_button.grid(row=17, column=0, sticky="w")
	vina_button.grid(row=17, column=1, sticky="w")
	gold_box_label = tk.Label(option_frame, text="GOLD Installation Folder:")
	gold_box_label.grid(row=18, column=0, sticky="w")
	gold_install_path = tk.StringVar()
	gold_install_box = ttk.Entry(option_frame, textvariable=gold_install_path)
	gold_install_box.grid(row=18, column=1, sticky="w")
	gold_install_button = ttk.Button(option_frame, text="Browse",
		command=lambda: get_directory(gold_install_path, gold_install_box))
	gold_install_button.grid(row=18, column=2, sticky="w")
	proteins_label = tk.Label(option_frame, text="Target Proteins:")
	proteins_label.grid(row=19, column=0, sticky="w")
	protein_selection = {"5IF3": tk.BooleanVar(), "1IEP": tk.BooleanVar(),
		"2W26": tk.BooleanVar(), "1ZYS": tk.BooleanVar(),
		"3RUK": tk.BooleanVar()}
	for i, protein in enumerate(list(protein_selection.keys())):
		protein_button = ttk.Checkbutton(option_frame,
			variable=protein_selection[protein], text=protein)
		protein_button.grid(row=20, column=i, sticky="w")

	# Custom docking section
	custom_gold_button_value = tk.BooleanVar()
	custom_gold_button = ttk.Checkbutton(option_frame,
		text="Custom GOLD Docking", variable=custom_gold_button_value,
		command=activate_docking_warning)
	custom_gold_button.grid(row=21, column=0, sticky="nw")
	conf_file_label = tk.Label(option_frame, text="gold.conf File:")
	conf_file_label.grid(row=22, column=0, sticky="nw")
	conf_filename = tk.StringVar()
	conf_file_box = ttk.Entry(option_frame, textvariable=conf_filename)
	conf_file_box.grid(row=22, column=1, sticky="nw")
	conf_file_button = ttk.Button(option_frame, text="Browse",
		command=lambda: get_filename(conf_filename, conf_file_box))
	conf_file_button.grid(row=22, column=2, sticky="nw")
	custom_vina_button_value = tk.BooleanVar()
	custom_vina_button = ttk.Checkbutton(option_frame,
		text="Custom Vina Docking", variable=custom_vina_button_value,
		command=activate_docking_warning)
	custom_vina_button.grid(row=23, column=0, sticky="nw")
	pdbqt_file_label = tk.Label(option_frame, text="Protein PDBQT File:")
	pdbqt_file_label.grid(row=24, column=0, sticky="nw")
	pdbqt_filename = tk.StringVar()
	pdbqt_file_box = ttk.Entry(option_frame, textvariable=pdbqt_filename)
	pdbqt_file_box.grid(row=24, column=1, sticky="nw")
	pdbqt_file_button = ttk.Button(option_frame, text="Browse",
		command=lambda: get_filename(pdbqt_filename, pdbqt_file_box))
	pdbqt_file_button.grid(row=24, column=2, sticky="nw")
	config_file_label = tk.Label(option_frame, text="Vina Config File:")
	config_file_label.grid(row=25, column=0, sticky="nw")
	config_filename = tk.StringVar()
	config_file_box = ttk.Entry(option_frame, textvariable=config_filename)
	config_file_box.grid(row=25, column=1, sticky="nw")
	config_file_button = ttk.Button(option_frame, text="Browse",
		command=lambda: get_filename(config_filename, config_file_box))
	config_file_button.grid(row=25, column=2, sticky="nw")

	assay_button = ttk.Button(option_frame, text="Assay Simulator",
		command=run_assay_simulator)
	assay_button.grid(row=26, column=0, sticky="nw")

	go_button = ttk.Button(option_frame, text="Go",
		command=handle_calculate_button)
	go_button.grid(row=27, column=0, sticky="nw")

	# Output textbox on right of GUI
	output_frame = tk.Frame()
	output_frame.grid(row=0, column=2, sticky="nsew", padx=FRAME_X_PADDING,
		pady=FRAME_Y_PADDING)
	output_label = tk.Label(output_frame, text="Calculation Output:")
	output_label.grid(row=0, column=0, sticky="nw")
	output_text = tk.Text(output_frame, state="disabled", width=TEXTBOX_WIDTH,
		height=TEXTBOX_HEIGHT)
	output_text.grid(row=1, column=0, sticky="nsew")
	output_frame.grid_columnconfigure(0, weight=1)
	output_frame.grid_rowconfigure(1, weight=1)

	window.columnconfigure((0, 1), weight=1)
	window.columnconfigure(2, weight=3)
	window.rowconfigure(0, weight=1)

	window.mainloop()
